#!/usr/bin/env python3
"""Compile the RCDAS SOP Database data layer.

Joins the document registry (source of truth for active/retired), the
structured work JSONs, and the glossary into web-ready files under data/.
Retired documents are never included. APET references are stripped (D-9).
"""
import json
import re
import unicodedata
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
SRC = Path("/Users/kristenhassen/Documents/Claude/Projects/Riverside County DAS/OUTPUTS/SOP Consolidation")
WORK = SRC / "_source-data" / "work"
OUT = HERE / "data"

# Consolidated documents that live at _source-data top level, keyed by the
# registry number or slug they correspond to.
TOP_LEVEL = {
    "200-23": SRC / "_source-data" / "200-23.json",
    "200-29": SRC / "_source-data" / "200-29.json",
    "iipp": SRC / "_source-data" / "iipp-consolidated.json",
    "facility-erp": SRC / "_source-data" / "facility-erp-consolidated.json",
    "fiscal-emergency": SRC / "_source-data" / "fiscal-emergency.json",
    "emergency-purchasing": SRC / "_source-data" / "emergency-purchasing.json",
}

# Registry titles whose work-file slug differs; registry-title-slug -> file stem.
ALIASES = {
    "bad-check-procedure-nsf-receipt-in-chameleon": "bad-check-procedure",
    "volunteer-guidelines-coachella-valley-and-san-jacinto-valley-animal-campuses": "volunteer-guidelines-cvac-sjvac",
    "volunteer-guidelines-riverside-animal-shelter": "volunteer-guidelines-riverside",
    "processing-dangerous-and-potentially-dangerous-restraint-orders-in-chameleon": "dangerous-restraint-orders",
    "cvac-animal-care-technician-act-duties": "cvac-act-duties",
    "completion-and-routing-of-the-rabies-control-investigation-report": "100-05-rabies-investigation-report",
}

# Documents flagged on BOTH active and sunsetted tracker tabs (handoff):
DUAL_LISTED = {"000-36", "000-46", "000-51", "000-58", "000-76", "needs-rescue"}

# Provenance, from the live "RCDAS Policies and Procedures" tracker (the same
# spreadsheet the team maintains on OneDrive; a copy lives beside the corpus as
# RCDAS_Live_Tracker.xlsx). Kristen's rule: a document was "updated by staff" if
# it has an entry in the New Draft column OR its Status is Done. Everything else
# was only reformatted into the 2026 template, not updated.
PROVENANCE = {
    "reviewed":    ("Updated by staff", "Staff wrote a new draft of this document during the 2026 consolidation."),
    "reformatted": ("Not yet updated", "Original content reformatted into the 2026 template. Staff have not written a new draft yet, so fees, contacts, and citations may be dated."),
    "reference":   ("Reference only", "Kept for reference; not slated for a rewrite."),
}
TRACKER_CATEGORY_SHEETS = [
    "Intake & Stray Processing", "Shelter Ops, Animal Care, Handl", "Shelter safety",
    "Outcomes & Placement", "Call Center & Customer Service", "Foster & Volunteer",
    "Administration & Personnel", "Technology & Records", "Veterinary & Medical",
    "Field Services",
]
_num_in = re.compile(r"(\b\d{3}-\d{1,3}|\bCC-\d{3})")
_prov_stop = set("sop draft rcdas policy procedure the for of and in a to new form word doc docx attachment".split())


def _tokset(s):
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().lower()
    s = re.sub(r"\d{4}-\d\d-\d\d|\d{6,8}", " ", s)
    return {w for w in re.findall(r"[a-z]+", s) if len(w) > 2 and w not in _prov_stop}


def load_tracker():
    """Read the live tracker's category tabs into match keys.
    Returns (by_num, fuzzy) where each value is a state string."""
    path = SRC / "RCDAS_Live_Tracker.xlsx"
    if not path.exists():
        return {}, []
    wb = openpyxl.load_workbook(path, data_only=True)
    by_num, fuzzy = {}, []
    rank = {"reviewed": 3, "reference": 2, "reformatted": 1}
    for sheet in TRACKER_CATEGORY_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        hdr = [c.value for c in ws[1]]
        ix = {h: i for i, h in enumerate(hdr) if h}
        ti = ix.get("Title", ix.get("duplicate"))
        si, ni = ix.get("Status"), ix.get("New Draft")
        if ti is None:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            if ti >= len(row) or not row[ti]:
                continue
            title = str(row[ti])
            status = str(row[si]).strip() if si is not None and si < len(row) and row[si] else ""
            nd = row[ni] if ni is not None and ni < len(row) else None
            has_nd = bool(nd and str(nd).strip() and str(nd).strip().lower() not in ("n/a", "na", "-"))
            if status.lower() == "reference only":
                state = "reference"
            elif has_nd or status.lower() == "done":
                state = "reviewed"          # updated by staff (new draft OR done)
            else:
                state = "reformatted"        # not yet updated
            m = _num_in.search(title)
            if m and rank[state] > rank.get(by_num.get(m.group(1)), 0):
                by_num[m.group(1)] = state
            ts = _tokset(title) | _tokset(nd or "")
            if len(ts) >= 2:
                fuzzy.append((ts, state))
    return by_num, fuzzy


def provenance_for(number, title, tracker):
    by_num, fuzzy = tracker
    state = None
    if number and number in by_num:
        state = by_num[number]
    if state is None:
        dt = _tokset(title)
        best, cand = 0.0, None
        for ts, st in fuzzy:
            j = len(dt & ts) / len(dt | ts) if (dt | ts) else 0
            if j > best:
                best, cand = j, st
        if best >= 0.45:
            state = cand
    if state is None:
        state = "reformatted"   # unmatched -> under-claim as not-yet-updated
    short, label = PROVENANCE[state]
    return {"state": state, "short": short, "label": label}

# SOPs most shelters have that RCDAS does not (staff-flagged gaps from the
# consolidation handoff, plus standard-of-care topics checked against the
# corpus: ASV Guidelines domains with no matching RCDAS document).
GAPS = [
    {"group": "Medical & Disease", "items": [
        {"topic": "Parvovirus treatment", "note": "A prevention procedure exists (puppy housing movement); there is no treatment protocol for infected animals."},
        {"topic": "Feline panleukopenia response", "note": "No document covers panleukopenia identification, isolation, or treatment."},
        {"topic": "Upper respiratory infection (URI) treatment", "note": "The most common shelter illness; no treatment or monitoring protocol."},
        {"topic": "Ringworm identification and treatment", "note": "No identification, lime-sulfur treatment, or clearance protocol."},
        {"topic": "Infectious disease outbreak response", "note": "Rabies quarantines are covered; there is no general outbreak response plan (closure criteria, testing, communication)."},
        {"topic": "Disease isolation protocols", "note": "Quarantine documents cover rabies observation only, not routine isolation for infectious disease."},
        {"topic": "Neonatal and fading kitten care", "note": "No protocol for bottle feeding, warmth, or fading kitten emergencies."},
    ]},
    {"group": "Operations & Population", "items": [
        {"topic": "Population management and capacity for care", "note": "No document ties intake, housing capacity, and daily population decisions together."},
        {"topic": "Daily population rounds", "note": "No SOP for structured daily walk-throughs checking every animal's status and pathway."},
        {"topic": "Community cat program", "note": "In progress per the consolidation; will define the TNR, CCP, and RTF outcome types."},
        {"topic": "Large-scale intake (hoarding and cruelty seizures)", "note": "No plan for intaking dozens of animals at once from a single case."},
        {"topic": "Whole-shelter sanitation", "note": "Kennel and area cleaning SOPs exist; no facility-wide sanitation program document."},
        {"topic": "Animal evacuation (shelter-wide)", "note": "No standalone evacuation SOP; the Facility ERP covers portions."},
        {"topic": "Missing animals", "note": "No SOP for animals missing from inventory."},
    ]},
    {"group": "Personnel & Administration", "items": [
        {"topic": "Euthanasia technician training and certification", "note": "No document covers training, certification, and competency requirements."},
        {"topic": "Phone-center triage (current practice)", "note": "Existing call center documents predate current triage practice."},
        {"topic": "Telework", "note": "No department telework policy."},
        {"topic": "Time Clock SOP", "note": "Policy 000-12 requires a Time Clock SOP that was never written."},
    ]},
]

NUM_RE = re.compile(r"^([0-9]{3}-[0-9]{1,3}[A-Za-z]?|CC-[0-9]{3})\b")


def slugify(s):
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")
    return s


def load_registry():
    wb = openpyxl.load_workbook(SRC / "RCDAS_Document_Registry.xlsx")
    ws = wb["Document Registry"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h) for h in rows[0]]
    out = []
    for r in rows[1:]:
        rec = dict(zip(hdr, r))
        if not rec.get("Document"):
            continue
        out.append(rec)
    return out


def index_json_files():
    """Map both number and subject-slug to each JSON path."""
    by_number, by_slug = {}, {}
    files = sorted(WORK.glob("*.json")) + [p for p in TOP_LEVEL.values() if p.exists()]
    for p in files:
        try:
            d = json.loads(p.read_text())
        except Exception as e:
            print(f"  WARN unreadable {p.name}: {e}")
            continue
        num = (d.get("number") or "").strip()
        if num:
            by_number.setdefault(num, (p, d))
        m = NUM_RE.match(p.stem)
        if m:
            by_number.setdefault(m.group(1), (p, d))
        subj = d.get("subject") or ""
        if subj:
            by_slug.setdefault(slugify(subj), (p, d))
        by_slug.setdefault(slugify(p.stem), (p, d))
    return by_number, by_slug


APET_RE = re.compile(r"\bAPET\b", re.IGNORECASE)


def strip_apet(obj):
    """Remove APET references anywhere in the structure (Decision D-9)."""
    if isinstance(obj, str):
        return obj
    if isinstance(obj, list):
        return [strip_apet(x) for x in obj if not (isinstance(x, str) and APET_RE.search(x))]
    if isinstance(obj, dict):
        return {k: strip_apet(v) for k, v in obj.items()}
    return obj


def doc_text(d):
    """Flatten a document's content for the search index."""
    parts = [d.get("purpose") or "", d.get("scope") or "", d.get("authority") or ""]
    for t in d.get("definitions") or []:
        parts.append(f"{t.get('term','')} {t.get('def','')}")
    def walk(sections):
        for s in sections or []:
            parts.append(s.get("heading") or "")
            for st in s.get("steps") or []:
                parts.append(st if isinstance(st, str) else json.dumps(st))
            walk(s.get("subsections"))
    walk(d.get("sections"))
    for r in d.get("references") or []:
        parts.append(r if isinstance(r, str) else json.dumps(r))
    return " ".join(p for p in parts if p)


def main():
    registry = load_registry()
    by_number, by_slug = index_json_files()
    tracker = load_tracker()

    active = [r for r in registry if str(r.get("Status", "")).startswith("Active")]
    retired_numbers = {str(r.get("Number") or "") for r in registry if not str(r.get("Status", "")).startswith("Active")}

    docs, unmatched = [], []
    for r in active:
        title = str(r["Document"]).strip()
        num = str(r.get("Number") or "").strip()
        cand = None
        if num and num != "-":
            cand = by_number.get(num)
        if cand is None:
            # strip a leading number from the title before slugging
            bare = NUM_RE.sub("", title).strip(" -")
            cand = by_slug.get(slugify(bare)) or by_slug.get(slugify(title))
        if cand is None:
            alias = ALIASES.get(slugify(NUM_RE.sub("", title).strip(" -")))
            if alias:
                cand = by_slug.get(slugify(alias))
        if cand is None and num in TOP_LEVEL:
            p = TOP_LEVEL[num]
            if p.exists():
                cand = (p, json.loads(p.read_text()))
        if cand is None:
            unmatched.append(title)
            continue
        p, d = cand
        d = strip_apet(d)
        docid = slugify(num if num and num != "-" else NUM_RE.sub("", title).strip(" -"))
        clean_num = num if num and num != "-" else None
        clean_title = NUM_RE.sub("", title).strip(" -")
        docs.append({
            "id": docid,
            "number": clean_num,
            "title": clean_title,
            # staff-facing type label: the registry says Policy/Protocol, but
            # the department's two document types are Policy and Procedure
            "type": "Procedure" if r.get("Type") == "Protocol" else r.get("Type"),
            "category": r.get("Category"),
            "status": "Draft pending approval",
            "provenance": provenance_for(clean_num, clean_title, tracker),
            "flag": r.get("Flag") or None,
            "dualListed": (num in DUAL_LISTED) or (slugify(title).find("needs-rescue") >= 0),
            "purpose": d.get("purpose"),
            "authority": d.get("authority"),
            "scope": d.get("scope"),
            "supersedes": d.get("supersedes"),
            "references": d.get("references") or [],
            "definitions": d.get("definitions") or [],
            "sections": d.get("sections") or [],
            "campusVariations": d.get("campusVariations") or [],
            "related": d.get("related") or [],
            "appendices": d.get("appendices") or [],
            "revisions": d.get("revisions") or [],
            "searchText": doc_text(d).lower(),
        })

    # Glossary: drop APET entries defensively; keep status for display.
    glossary = json.loads((SRC / "Glossary" / "glossary.json").read_text())
    glossary = [g for g in glossary if not APET_RE.search(json.dumps(g))]

    # Merge terms extracted from the staff reference docs (parse_docs.py),
    # skipping anything the master glossary already covers.
    extra_path = OUT / "extra_terms.json"
    if extra_path.exists():
        have = {re.sub(r"[^A-Z0-9]", "", g["term"].upper()) for g in glossary}
        extras = [t for t in json.loads(extra_path.read_text())
                  if re.sub(r"[^A-Z0-9]", "", t["term"].upper()) not in have]
        glossary += extras
        print(f"merged {len(extras)} extra terms from staff reference docs")

    OUT.mkdir(exist_ok=True)
    (OUT / "corpus.json").write_text(json.dumps({
        "generated": "build_corpus.py",
        "counts": {"documents": len(docs),
                   "policies": sum(1 for d in docs if d["type"] == "Policy"),
                   "procedures": sum(1 for d in docs if d["type"] == "Procedure")},
        "documents": docs,
    }, ensure_ascii=False))
    (OUT / "glossary.json").write_text(json.dumps(glossary, ensure_ascii=False))
    (OUT / "gaps.json").write_text(json.dumps(GAPS, ensure_ascii=False, indent=1))

    print(f"registry active: {len(active)}  matched: {len(docs)}  unmatched: {len(unmatched)}")
    for t in unmatched:
        print("  UNMATCHED:", t)
    print(f"glossary terms: {len(glossary)}")
    kb = (OUT / 'corpus.json').stat().st_size // 1024
    print(f"corpus.json: {kb} KB")


if __name__ == "__main__":
    main()
